"""
routes/inquiries.py â€” Inquiries (HeavyLift CRM)
Features: auto-followup notification, optional refs, offer linkage, WhatsApp, Excel export
"""
from datetime import date, timedelta
import io
import os
import urllib.parse
from pathlib import Path
from uuid import uuid4

from flask import Blueprint, abort, current_app, flash, jsonify, make_response, redirect, render_template, request, send_file, session, url_for
from werkzeug.utils import secure_filename

from ..database import close_db, get_db
from .inquiry_helpers import (
    apply_inquiry_filters,
    build_inquiry_scope,
    calculate_total_fees,
    DOCUMENT_TYPE_LABELS,
    fetch_inquiry,
    group_inquiry_documents,
    load_form_options,
    load_inquiry_documents,
    load_index_lookups,
    parse_amount,
    parse_date,
    parse_index_filters,
    parse_sort_args,
    render_inquiry_form,
    validate_inquiry_form,
    validate_teacher_form_access,
)
from .auth import login_required, role_required
from ..webservices.notifications import create_notification
from ..validation import clean_choice, clean_optional_text, parse_optional_int

inquiries_bp = Blueprint("inquiries", __name__, url_prefix="/inquiries")

ALLOWED_UPLOAD_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf"}
PHOTO_DOCUMENT_TYPES = {"student_photo"}
ATTACHMENT_TYPES = {"student_photo", "govt_id"}
ATTACHMENT_COLUMN_MAP = {
    "student_photo": {
        "file_path": "student_photo_file_path",
        "original_filename": "student_photo_original_filename",
        "stored_filename": "student_photo_stored_filename",
        "mime_type": "student_photo_mime_type",
        "file_size": "student_photo_file_size",
    },
    "govt_id": {
        "file_path": "govt_id_file_path",
        "original_filename": "govt_id_original_filename",
        "stored_filename": "govt_id_stored_filename",
        "mime_type": "govt_id_mime_type",
        "file_size": "govt_id_file_size",
    },
}


def _uploads_root():
    base = Path(current_app.root_path)
    folder = Path(current_app.config.get("UPLOAD_FOLDER", "uploads"))
    return folder if folder.is_absolute() else base / folder


def _inquiry_documents_dir(iid):
    target = _uploads_root() / "inquiries" / str(iid)
    target.mkdir(parents=True, exist_ok=True)
    return target


def _is_allowed_upload(file_storage):
    filename = secure_filename(file_storage.filename or "")
    if not filename:
      return False
    ext = Path(filename).suffix.lower()
    if ext in ALLOWED_UPLOAD_EXTENSIONS:
        return True
    mime = (file_storage.mimetype or "").lower()
    return mime.startswith("image/") or mime == "application/pdf"


def _normalize_upload_label(label):
    return (label or "").strip()


def _collect_attachment_uploads(files):
    uploads = []
    for attachment_type in ATTACHMENT_TYPES:
        upload = files.get(attachment_type)
        if upload and upload.filename:
            uploads.append((attachment_type, upload))
    return uploads


def _collect_supporting_uploads(files):
    uploads = []
    for upload in files.getlist("supporting_documents"):
        if upload and upload.filename:
            uploads.append(("supporting_document", upload))
    return uploads


def _save_inquiry_upload(iid, document_type, upload):
    filename = secure_filename(upload.filename or "")
    if not filename:
        raise ValueError("Please choose a valid file to upload.")
    if not _is_allowed_upload(upload):
        raise ValueError(f"{DOCUMENT_TYPE_LABELS.get(document_type, 'Document')} must be an image or PDF file.")

    suffix = Path(filename).suffix.lower() or ".bin"
    stored_name = f"{document_type}_{uuid4().hex}{suffix}"
    docs_dir = _inquiry_documents_dir(iid)
    file_path = docs_dir / stored_name
    upload.save(file_path)
    return {
        "original_filename": filename,
        "stored_filename": stored_name,
        "file_path": str(file_path),
        "mime_type": upload.mimetype or None,
        "file_size": os.path.getsize(file_path),
    }


def _save_inquiry_attachment(iid, attachment_type, upload):
    if attachment_type not in ATTACHMENT_TYPES:
        raise ValueError("Unsupported attachment type.")
    return _save_inquiry_upload(iid, attachment_type, upload)


def _load_document_rows(cur, iid):
    cur.execute("SELECT * FROM inquiry_documents WHERE inquiry_id=%s ORDER BY created_at ASC, id ASC;", (iid,))
    return cur.fetchall()


def _split_document_groups(documents):
    grouped = {key: [] for key in DOCUMENT_TYPE_LABELS}
    for document in documents or []:
        document_type = document["document_type"]
        if document_type in {"govt_id_front", "govt_id_back"}:
            document_type = "govt_id"
        grouped.setdefault(document_type, []).append(document)
    return grouped


def _remove_document_file(file_path):
    try:
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
    except OSError:
        pass


def _missing_attachment_response(label, as_download=False):
    placeholder_svg = f"""<svg xmlns='http://www.w3.org/2000/svg' width='900' height='520' viewBox='0 0 900 520'>
  <rect width='900' height='520' rx='24' fill='#f8fafc'/>
  <rect x='24' y='24' width='852' height='472' rx='20' fill='#ffffff' stroke='#e5e7eb'/>
  <text x='450' y='240' text-anchor='middle' font-family='Arial, sans-serif' font-size='34' fill='#1f2937'>{label}</text>
  <text x='450' y='288' text-anchor='middle' font-family='Arial, sans-serif' font-size='18' fill='#6b7280'>Attachment file is unavailable</text>
</svg>"""
    if as_download:
        return make_response("Attachment file is unavailable.", 404)
    response = make_response(placeholder_svg)
    response.headers["Content-Type"] = "image/svg+xml"
    response.headers["Cache-Control"] = "no-store"
    return response


def _replace_single_document(cur, iid, document_type):
    document_types = [document_type]
    if document_type == "govt_id":
        document_types = list(LEGACY_GOVT_ID_TYPES)
    placeholders = ",".join(["%s"] * len(document_types))
    cur.execute(
        f"SELECT id, file_path FROM inquiry_documents WHERE inquiry_id=%s AND document_type IN ({placeholders}) ORDER BY created_at DESC, id DESC;",
        [iid, *document_types],
    )
    return cur.fetchall()


def _persist_inquiry_attachments(cur, iid, uploads, existing_inquiry=None):
    removed_files = []
    saved_files = []
    try:
        for attachment_type, upload in uploads:
            saved = _save_inquiry_attachment(iid, attachment_type, upload)
            saved_files.append(saved["file_path"])
            column_map = ATTACHMENT_COLUMN_MAP[attachment_type]
            cur.execute(
                f"""
                UPDATE inquiries SET
                  {column_map['file_path']}=%s,
                  {column_map['original_filename']}=%s,
                  {column_map['stored_filename']}=%s,
                  {column_map['mime_type']}=%s,
                  {column_map['file_size']}=%s
                WHERE id=%s;
                """,
                (
                    saved["file_path"],
                    saved["original_filename"],
                    saved["stored_filename"],
                    saved["mime_type"],
                    saved["file_size"],
                    iid,
                ),
            )
            old_path = None
            if existing_inquiry:
                old_path = existing_inquiry.get(column_map["file_path"])
            if old_path and old_path != saved["file_path"]:
                removed_files.append(old_path)
    except Exception:
        for file_path in saved_files:
            _remove_document_file(file_path)
        raise

    return removed_files, saved_files


def _persist_inquiry_documents(cur, iid, uploads, replace_single=True):
    removed_files = []
    saved_files = []
    try:
        if replace_single:
            for document_type, _upload in uploads:
                if document_type in SINGLE_DOCUMENT_TYPES:
                    existing = _replace_single_document(cur, iid, document_type)
                    for row in existing:
                        cur.execute("DELETE FROM inquiry_documents WHERE id=%s;", (row["id"],))
                        removed_files.append(row.get("file_path"))

        for document_type, upload in uploads:
            saved = _save_inquiry_upload(iid, document_type, upload)
            saved_files.append(saved["file_path"])
            cur.execute(
                """
                INSERT INTO inquiry_documents
                  (inquiry_id, document_type, original_filename, stored_filename, file_path, mime_type, file_size)
                VALUES (%s,%s,%s,%s,%s,%s,%s);
                """,
                (
                    iid,
                    document_type,
                    saved["original_filename"],
                    saved["stored_filename"],
                    saved["file_path"],
                    saved["mime_type"],
                    saved["file_size"],
                ),
            )
    except Exception:
        for file_path in saved_files:
            _remove_document_file(file_path)
        raise

    return removed_files, saved_files


@inquiries_bp.route("/")
@login_required
def index():
    role = session.get("role")
    loc_id = session.get("location_id")
    filters = parse_index_filters(request.args)
    sort_col, sort_dir = parse_sort_args(request.args)

    base, params = build_inquiry_scope(role, loc_id)
    base, params, warnings = apply_inquiry_filters(base, params, filters)
    for warning in warnings:
        flash(warning, "warning")

    conn = get_db()
    cur = conn.cursor()
    cur.execute(f"SELECT i.*,l.name AS location_name,c.name AS course_name,o.name AS offer_name {base} ORDER BY {sort_col} {sort_dir};", params)
    inquiries = cur.fetchall()
    locations, courses = load_index_lookups(cur, role, loc_id)
    close_db(conn, commit=False)
    return render_template(
        "inquiries/index.html",
        inquiries=inquiries,
        locations=locations,
        courses=courses,
        filters=filters,
        sort_col=sort_col,
        sort_dir=sort_dir,
        today=date.today(),
    )


@inquiries_bp.route("/add", methods=["GET", "POST"])
@login_required
@role_required("admin", "developer")
def add():
    role = session.get("role")
    assigned_loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    locs, courses_list, offers = load_form_options(cur, role, assigned_loc_id)
    close_db(conn, commit=False)

    defaults = {
        "inquiry_date": date.today().isoformat(),
        "followup_date": (date.today() + timedelta(days=10)).isoformat(),
    }

    if request.method == "POST":
        form = request.form
        try:
            location_id = parse_optional_int(form.get("location_id"), "Location")
            if role == "teacher" and assigned_loc_id:
                location_id = assigned_loc_id
            course_id = parse_optional_int(form.get("course_id"), "Course")
            offer_id = parse_optional_int(form.get("offer_id"), "Offer")
            fees_total = calculate_total_fees(course_id, offer_id)
            cleaned = validate_inquiry_form(form, fees_total)
            attachment_uploads = _collect_attachment_uploads(request.files)
            supporting_uploads = _collect_supporting_uploads(request.files)
            uploads = supporting_uploads
            for document_type, upload in attachment_uploads + supporting_uploads:
                if not _is_allowed_upload(upload):
                    raise ValueError(f"{DOCUMENT_TYPE_LABELS.get(document_type, 'Document')} must be an image or PDF file.")
            if role == "teacher" and assigned_loc_id:
                conn_check = get_db()
                cur_check = conn_check.cursor()
                try:
                    validate_teacher_form_access(cur_check, course_id, offer_id, assigned_loc_id)
                finally:
                    close_db(conn_check, commit=False)
            conn = get_db()
            cur = conn.cursor()
            try:
                cur.execute(
                    """
                    INSERT INTO inquiries
                      (name,gender,mobile,location_id,city,state,course_id,offer_id,
                       inquiry_date,followup_date,admission_date,status,fees_total,fees_paid,
                                                     ref1_name,ref1_type,ref1_mobile,ref1_amount_paid,
                                                     ref2_name,ref2_type,ref2_mobile,ref2_amount_paid,
                                                     ref3_name,ref3_type,ref3_mobile,ref3_amount_paid,
                             emergency1_name,emergency1_mobile,emergency1_relation,
                             emergency2_name,emergency2_mobile,emergency2_relation,
                             emergency3_name,emergency3_mobile,emergency3_relation)
                                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    RETURNING id;
                    """,
                    (
                        cleaned["name"],
                        cleaned["gender"],
                        cleaned["mobile"],
                        location_id,
                        cleaned["city"],
                        cleaned["state"],
                        course_id,
                        offer_id,
                        cleaned["inquiry_date"],
                        cleaned["followup_date"],
                        cleaned["admission_date"],
                        cleaned["status"],
                        fees_total,
                        cleaned["fees_paid"],
                        cleaned["ref1_name"],
                        cleaned["ref1_type"],
                        cleaned["ref1_mobile"],
                        cleaned["ref1_amount_paid"],
                        cleaned["ref2_name"],
                        cleaned["ref2_type"],
                        cleaned["ref2_mobile"],
                        cleaned["ref2_amount_paid"],
                        cleaned["ref3_name"],
                        cleaned["ref3_type"],
                        cleaned["ref3_mobile"],
                        cleaned["ref3_amount_paid"],
                        cleaned["emergency1_name"],
                        cleaned["emergency1_mobile"],
                        cleaned["emergency1_relation"],
                        cleaned["emergency2_name"],
                        cleaned["emergency2_mobile"],
                        cleaned["emergency2_relation"],
                        cleaned["emergency3_name"],
                        cleaned["emergency3_mobile"],
                        cleaned["emergency3_relation"],
                    ),
                )
                inquiry_row = cur.fetchone() or {}
                inquiry_id = inquiry_row.get("id")
                removed_files = []
                if attachment_uploads and inquiry_id:
                    attachment_removed_files, _saved_files = _persist_inquiry_attachments(cur, inquiry_id, attachment_uploads)
                    removed_files.extend(attachment_removed_files)
                if uploads and inquiry_id:
                    supporting_removed_files, _saved_files = _persist_inquiry_documents(cur, inquiry_id, uploads, replace_single=False)
                    removed_files.extend(supporting_removed_files)
                close_db(conn)
                for file_path in removed_files:
                    _remove_document_file(file_path)
            except Exception:
                close_db(conn, commit=False)
                raise
            create_notification(
                f"Follow-up due: {cleaned['name']}",
                f"Follow-up scheduled on {cleaned['followup_date'] or 'N/A'} for {cleaned['name']} ({cleaned['mobile']}).",
                target_role="admin",
            )
            flash("Inquiry added.", "success")
            return redirect(url_for("inquiries.index"))
        except ValueError as exc:
            flash(str(exc), "danger")
            return render_inquiry_form(
                inquiry=None,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults=defaults,
                action="Add",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=str(exc),
                documents_by_type={},
            )
        except Exception:
            current_app.logger.exception("Failed to add inquiry")
            message = "Unable to save inquiry right now."
            flash(message, "danger")
            return render_inquiry_form(
                inquiry=None,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults=defaults,
                action="Add",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=message,
                documents_by_type={},
            )

    return render_inquiry_form(
        inquiry=None,
        locations=locs,
        courses=courses_list,
        offers=offers,
        defaults=defaults,
        action="Add",
        documents_by_type={},
    )


@inquiries_bp.route("/<int:iid>/edit", methods=["GET", "POST"])
@login_required
@role_required("admin", "developer", "teacher")
def edit(iid):
    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id)
    locs, courses_list, offers = load_form_options(cur, role, loc_id)
    document_rows = _load_document_rows(cur, iid)
    documents_by_type = _split_document_groups(document_rows)
    close_db(conn, commit=False)
    if not inquiry:
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))

    if request.method == "POST":
        form = request.form
        try:
            location_id = parse_optional_int(form.get("location_id"), "Location")
            course_id = parse_optional_int(form.get("course_id"), "Course")
            offer_id = parse_optional_int(form.get("offer_id"), "Offer")
            fees_total = calculate_total_fees(course_id, offer_id)
            cleaned = validate_inquiry_form(form, fees_total)
            attachment_uploads = _collect_attachment_uploads(request.files)
            supporting_uploads = _collect_supporting_uploads(request.files)
            uploads = supporting_uploads
            for document_type, upload in attachment_uploads + supporting_uploads:
                if not _is_allowed_upload(upload):
                    raise ValueError(f"{DOCUMENT_TYPE_LABELS.get(document_type, 'Document')} must be an image or PDF file.")
            conn = get_db()
            cur = conn.cursor()
            current_inquiry = fetch_inquiry(cur, iid, role, loc_id)
            if not current_inquiry:
                close_db(conn, commit=False)
                flash("Not found.", "danger")
                return redirect(url_for("inquiries.index"))
            if role == "teacher" and loc_id:
                try:
                    validate_teacher_form_access(cur, course_id, offer_id, loc_id)
                except ValueError:
                    close_db(conn, commit=False)
                    raise
            cur.execute(
                """
                UPDATE inquiries SET
                  name=%s,gender=%s,mobile=%s,location_id=%s,city=%s,state=%s,
                  course_id=%s,offer_id=%s,inquiry_date=%s,followup_date=%s,
                  admission_date=%s,status=%s,fees_total=%s,fees_paid=%s,
                                                                        ref1_name=%s,ref1_type=%s,ref1_mobile=%s,ref1_amount_paid=%s,
                                                                        ref2_name=%s,ref2_type=%s,ref2_mobile=%s,ref2_amount_paid=%s,
                                                                        ref3_name=%s,ref3_type=%s,ref3_mobile=%s,ref3_amount_paid=%s,
                                                                        emergency1_name=%s,emergency1_mobile=%s,emergency1_relation=%s,
                                                                        emergency2_name=%s,emergency2_mobile=%s,emergency2_relation=%s,
                                                                        emergency3_name=%s,emergency3_mobile=%s,emergency3_relation=%s
                WHERE id=%s;
                """,
                (
                    cleaned["name"],
                    cleaned["gender"],
                    cleaned["mobile"],
                    loc_id if role == "teacher" and loc_id else location_id,
                    cleaned["city"],
                    cleaned["state"],
                    course_id,
                    offer_id,
                    cleaned["inquiry_date"],
                    cleaned["followup_date"],
                    cleaned["admission_date"],
                    cleaned["status"],
                    fees_total,
                    cleaned["fees_paid"],
                    cleaned["ref1_name"],
                    cleaned["ref1_type"],
                    cleaned["ref1_mobile"],
                    cleaned["ref1_amount_paid"],
                    cleaned["ref2_name"],
                    cleaned["ref2_type"],
                    cleaned["ref2_mobile"],
                    cleaned["ref2_amount_paid"],
                    cleaned["ref3_name"],
                    cleaned["ref3_type"],
                    cleaned["ref3_mobile"],
                    cleaned["ref3_amount_paid"],
                    cleaned["emergency1_name"],
                    cleaned["emergency1_mobile"],
                    cleaned["emergency1_relation"],
                    cleaned["emergency2_name"],
                    cleaned["emergency2_mobile"],
                    cleaned["emergency2_relation"],
                    cleaned["emergency3_name"],
                    cleaned["emergency3_mobile"],
                    cleaned["emergency3_relation"],
                    iid,
                ),
            )
            removed_files = []
            if attachment_uploads:
                attachment_removed_files, _saved_files = _persist_inquiry_attachments(cur, iid, attachment_uploads, existing_inquiry=current_inquiry)
                removed_files.extend(attachment_removed_files)
            if uploads:
                supporting_removed_files, _saved_files = _persist_inquiry_documents(cur, iid, uploads, replace_single=True)
                removed_files.extend(supporting_removed_files)
            close_db(conn)
            for file_path in removed_files:
                _remove_document_file(file_path)
            flash("Updated.", "success")
            return redirect(url_for("inquiries.index"))
        except ValueError as exc:
            flash(str(exc), "danger")
            return render_inquiry_form(
                inquiry=inquiry,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults={},
                action="Edit",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=str(exc),
                documents_by_type=documents_by_type,
            )
        except Exception:
            current_app.logger.exception("Failed to update inquiry %s", iid)
            message = "Unable to update inquiry right now."
            flash(message, "danger")
            return render_inquiry_form(
                inquiry=inquiry,
                locations=locs,
                courses=courses_list,
                offers=offers,
                defaults={},
                action="Edit",
                form_data=request.form.to_dict(flat=True),
                form_error_popup=message,
                documents_by_type=documents_by_type,
            )

    return render_inquiry_form(
        inquiry=inquiry,
        locations=locs,
        courses=courses_list,
        offers=offers,
        defaults={},
        action="Edit",
        documents_by_type=documents_by_type,
    )


@inquiries_bp.route("/<int:iid>/documents/<int:doc_id>")
@login_required
def document_file(iid, doc_id):
    role = session.get("role")
    loc_id = session.get("location_id")
    download = request.args.get("download") == "1"
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id)
    if not inquiry:
        close_db(conn, commit=False)
        abort(404)

    cur.execute(
        "SELECT * FROM inquiry_documents WHERE id=%s AND inquiry_id=%s;",
        (doc_id, iid),
    )
    document = cur.fetchone()
    close_db(conn, commit=False)
    if not document:
        abort(404)

    file_path = document.get("file_path")
    if not file_path or not os.path.exists(file_path):
        return _missing_attachment_response(DOCUMENT_TYPE_LABELS.get(document.get("document_type"), "Attachment"), as_download=download)

    return send_file(
        file_path,
        mimetype=document.get("mime_type") or None,
        as_attachment=download,
        download_name=document.get("original_filename") or Path(file_path).name,
        conditional=True,
    )


@inquiries_bp.route("/<int:iid>/attachments/<attachment_type>")
@login_required
def attachment_file(iid, attachment_type):
    if attachment_type not in ATTACHMENT_COLUMN_MAP:
        abort(404)

    role = session.get("role")
    loc_id = session.get("location_id")
    download = request.args.get("download") == "1"
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id)
    close_db(conn, commit=False)
    if not inquiry:
        abort(404)

    column_map = ATTACHMENT_COLUMN_MAP[attachment_type]
    file_path = inquiry.get(column_map["file_path"])
    if not file_path or not os.path.exists(file_path):
        return _missing_attachment_response(DOCUMENT_TYPE_LABELS.get(attachment_type, "Attachment"), as_download=download)

    return send_file(
        file_path,
        mimetype=inquiry.get(column_map["mime_type"]) or None,
        as_attachment=download,
        download_name=inquiry.get(column_map["original_filename"]) or Path(file_path).name,
        conditional=True,
    )


@inquiries_bp.route("/<int:iid>/documents/<int:doc_id>/delete", methods=["POST"])
@login_required
@role_required("admin", "developer", "teacher")
def delete_document(iid, doc_id):
    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id)
    if not inquiry:
        close_db(conn, commit=False)
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))

    cur.execute(
        "SELECT id, file_path FROM inquiry_documents WHERE id=%s AND inquiry_id=%s;",
        (doc_id, iid),
    )
    document = cur.fetchone()
    if not document:
        close_db(conn, commit=False)
        flash("Document not found.", "danger")
        return redirect(url_for("inquiries.edit", iid=iid))

    cur.execute("DELETE FROM inquiry_documents WHERE id=%s AND inquiry_id=%s;", (doc_id, iid))
    close_db(conn)
    _remove_document_file(document.get("file_path"))
    flash("Document deleted.", "success")
    return redirect(url_for("inquiries.edit", iid=iid))


@inquiries_bp.route("/<int:iid>/attachments/<attachment_type>/delete", methods=["POST"])
@login_required
@role_required("admin", "developer", "teacher")
def delete_attachment(iid, attachment_type):
    if attachment_type not in ATTACHMENT_COLUMN_MAP:
        abort(404)

    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id)
    if not inquiry:
        close_db(conn, commit=False)
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))

    column_map = ATTACHMENT_COLUMN_MAP[attachment_type]
    file_path = inquiry.get(column_map["file_path"])
    cur.execute(
        f"UPDATE inquiries SET {column_map['file_path']}=NULL, {column_map['original_filename']}=NULL, {column_map['stored_filename']}=NULL, {column_map['mime_type']}=NULL, {column_map['file_size']}=NULL WHERE id=%s;",
        (iid,),
    )
    close_db(conn)
    _remove_document_file(file_path)
    flash(f"{DOCUMENT_TYPE_LABELS.get(attachment_type, 'Attachment')} deleted.", "success")
    return redirect(url_for("inquiries.edit", iid=iid))


@inquiries_bp.route("/<int:iid>/delete", methods=["POST"])
@login_required
@role_required("admin", "developer")
def delete(iid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT student_photo_file_path, govt_id_file_path FROM inquiries WHERE id=%s;",
        (iid,),
    )
    inquiry = cur.fetchone() or {}
    cur.execute("SELECT file_path FROM inquiry_documents WHERE inquiry_id=%s;", (iid,))
    document_paths = [row.get("file_path") for row in cur.fetchall()]
    cur.execute("DELETE FROM inquiries WHERE id=%s;", (iid,))
    close_db(conn)
    for file_path in [inquiry.get("student_photo_file_path"), inquiry.get("govt_id_file_path"), *document_paths]:
        _remove_document_file(file_path)
    flash("Deleted.", "success")
    return redirect(url_for("inquiries.index"))


@inquiries_bp.route("/<int:iid>/convert", methods=["POST"])
@login_required
@role_required("admin", "developer", "teacher")
def convert(iid):
    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id)
    if not inquiry:
        close_db(conn, commit=False)
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))
    cur.execute(
        "UPDATE inquiries SET status='Converted', admission_date=COALESCE(admission_date,CURRENT_DATE) WHERE id=%s;",
        (iid,),
    )
    close_db(conn)
    create_notification(f"Admission: {inquiry['name']}", f"{inquiry['name']} has been converted to a student.", "admin")
    flash(f"{inquiry['name']} converted to student.", "success")
    return redirect(url_for("inquiries.index"))


@inquiries_bp.route("/<int:iid>/followup", methods=["GET", "POST"])
@login_required
@role_required("admin", "developer", "teacher")
def followup(iid):
    role = session.get("role")
    loc_id = session.get("location_id")
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id, with_joins=True)
    if not inquiry:
        close_db(conn, commit=False)
        flash("Not found.", "danger")
        return redirect(url_for("inquiries.index"))

    if request.method == "POST":
        try:
            conversation = clean_optional_text(request.form.get("conversation", ""), "Conversation", max_length=4000, multiline=True)
            followup_date = parse_date(request.form.get("followup_date"), "Follow-up date")
            status = clean_choice(request.form.get("status", inquiry["status"]), "Status", {"Open", "Converted", "Closed"})
            admission_date = parse_date(request.form.get("admission_date"), "Admission date")
            if followup_date and followup_date < inquiry["inquiry_date"]:
                raise ValueError("Follow-up date cannot be earlier than inquiry date.")
            if admission_date and admission_date < inquiry["inquiry_date"]:
                raise ValueError("Admission date cannot be earlier than inquiry date.")
            cur.execute(
                "INSERT INTO followups (inquiry_id,conversation,followup_date,status) VALUES (%s,%s,%s,%s);",
                (iid, conversation, followup_date.isoformat() if followup_date else None, status),
            )
            cur.execute(
                "UPDATE inquiries SET status=%s,followup_date=%s,admission_date=COALESCE(%s::date,admission_date) WHERE id=%s;",
                (status, followup_date.isoformat() if followup_date else None, admission_date.isoformat() if admission_date else None, iid),
            )
            close_db(conn)
            if followup_date:
                create_notification(f"Next follow-up: {inquiry['name']}", f"Scheduled for {followup_date.isoformat()}.", "admin")
            flash("Follow-up saved.", "success")
            return redirect(url_for("inquiries.followup", iid=iid))
        except ValueError as exc:
            flash(str(exc), "danger")
        except Exception:
            current_app.logger.exception("Failed to save follow-up for inquiry %s", iid)
            flash("Unable to save follow-up right now.", "danger")

    cur.execute("SELECT * FROM followups WHERE inquiry_id=%s ORDER BY created_at DESC;", (iid,))
    followups = cur.fetchall()
    close_db(conn, commit=False)
    default_next = (date.today() + timedelta(days=7)).isoformat()
    return render_template("inquiries/followup.html", inquiry=inquiry, followups=followups, default_next=default_next)


@inquiries_bp.route("/<int:iid>/whatsapp-send", methods=["POST"])
@login_required
def send_whatsapp(iid):
    """Return wa.me link for direct WhatsApp send (or call API if configured)."""
    if not request.is_json:
        return jsonify({"ok": False, "msg": "JSON body required"}), 400
    role = session.get("role")
    loc_id = session.get("location_id")
    payload = request.get_json(silent=True) or {}
    conn = get_db()
    cur = conn.cursor()
    inquiry = fetch_inquiry(cur, iid, role, loc_id)
    if not inquiry:
        close_db(conn, commit=False)
        return jsonify({"ok": False, "msg": "Not found"}), 404

    try:
        msg_id = parse_optional_int(payload.get("msg_id"), "Message template")
        msg_text = clean_optional_text(payload.get("message"), "Message", max_length=2000, multiline=True) or ""
    except ValueError as exc:
        close_db(conn, commit=False)
        return jsonify({"ok": False, "msg": str(exc)}), 400
    if msg_id:
        cur.execute("SELECT description FROM whatsapp_msgs WHERE id=%s;", (msg_id,))
        template = cur.fetchone()
        if template:
            msg_text = (template["description"] or "").replace("[NAME]", inquiry["name"]).replace("[MOBILE]", inquiry["mobile"])
    close_db(conn, commit=False)

    mobile = inquiry["mobile"].replace(" ", "").replace("-", "").replace("+", "")
    if not mobile.startswith("91"):
        mobile = "91" + mobile
    wa_url = f"https://wa.me/{mobile}?text={urllib.parse.quote(msg_text)}"
    return jsonify({"ok": True, "url": wa_url})


@inquiries_bp.route("/export")
@login_required
@role_required("admin", "developer", "teacher")
def export():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    role = session.get("role")
    loc_id = session.get("location_id")
    base, params = build_inquiry_scope(role, loc_id)
    conn = get_db()
    cur = conn.cursor()
    cur.execute(f"SELECT i.*,l.name AS location_name,c.name AS course_name {base} ORDER BY i.inquiry_date DESC;", params)
    rows = cur.fetchall()
    close_db(conn, commit=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inquiries"
    headers = [
        "ID", "Name", "Gender", "Mobile", "Location", "City", "State", "Course",
        "Inquiry Date", "Followup Date", "Admission Date", "Status",
        "Fees Total", "Fees Paid", "Pending", "Ref1 Name", "Ref1 Mobile",
    ]
    hfill = PatternFill("solid", fgColor="F59E0B")
    hfont = Font(color="000000", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(rows, 2):
        pending = float(row.get("fees_total") or 0) - float(row.get("fees_paid") or 0)
        values = [
            row["id"], row["name"], row["gender"], row["mobile"], row["location_name"],
            row["city"], row["state"], row["course_name"],
            str(row["inquiry_date"]) if row["inquiry_date"] else "",
            str(row["followup_date"]) if row["followup_date"] else "",
            str(row["admission_date"]) if row["admission_date"] else "",
            row["status"], row.get("fees_total", 0), row.get("fees_paid", 0),
            pending, row.get("ref1_name", ""), row.get("ref1_mobile", ""),
        ]
        for col_index, value in enumerate(values, 1):
            ws.cell(row=row_index, column=col_index, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = "attachment; filename=inquiries.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response
